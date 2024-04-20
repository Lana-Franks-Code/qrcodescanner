from flask import Flask, render_template, request
from openpyxl import load_workbook, Workbook  # Added Workbook import
from datetime import datetime
import cv2
from pyzbar.pyzbar import decode
import time
import threading
import os

app = Flask(__name__)

# Global variables
scanning_active = False
message = ""

# Function to update attendance in Excel sheet
def update_attendance(name):
    try:
        wb = load_workbook("attendance.xlsx")
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value == name:
            return "Already Logged"
    next_row = ws.max_row + 1
    ws[f"A{next_row}"] = name
    ws[f"B{next_row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb.save("attendance.xlsx")
    return f"Welcome {name}"

# Function to continuously scan for QR codes
# Function to continuously scan for QR codes
def scan_qr():
    global scanning_active, message
    cap = cv2.VideoCapture(0)
    while scanning_active:
        ret, frame = cap.read()
        decoded_objects = decode(frame)
        if decoded_objects:
            name = decoded_objects[0].data.decode('utf-8')
            message = update_attendance(name)
            print(f"Decoded QR code: {name}")
            print(f"Attendance update message: {message}")
            time.sleep(2)  # Delay between scans
        else:
            message = "No QR code detected"
            print(message)
        time.sleep(0.1)  # Optional: Add a small delay to reduce CPU usage

# Flask routes
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/start_scan")
def start_scan():
    global scanning_active
    scanning_active = True
    thread = threading.Thread(target=scan_qr)
    thread.daemon = True
    thread.start()
    return "Scanning started"

@app.route("/stop_scan")
def stop_scan():
    global scanning_active
    scanning_active = False
    return "Scanning stopped"

@app.route("/get_message")
def get_message():
    global message
    return message

if __name__ == "__main__":
    app.run(debug=True)
